import google.generativeai as genai
from django.conf import settings
from PyPDF2 import PdfReader
from docx import Document as DocxDocument
import io
import openpyxl
import xlrd  # For reading .xls files

# Configure Gemini
MODEL_NAME = "gemini-2.0-flash"
model = genai.GenerativeModel(MODEL_NAME)

def read_document_content(uploaded_file):
    """
    Read content from uploaded file based on file type
    """
    try:
        # Reset file pointer to beginning
        uploaded_file.seek(0)
       
        file_name = uploaded_file.name.lower()
        content = ""
       
        if file_name.endswith('.txt'):
            content = uploaded_file.read().decode('utf-8')
       
        elif file_name.endswith('.pdf'):
            # Create a BytesIO object from the uploaded file
            pdf_file = io.BytesIO(uploaded_file.read())
            pdf_reader = PdfReader(pdf_file)
           
            for page in pdf_reader.pages:
                content += page.extract_text() + "\n"
       
        elif file_name.endswith('.docx'):
            # Create a BytesIO object from the uploaded file
            docx_file = io.BytesIO(uploaded_file.read())
            doc = DocxDocument(docx_file)
           
            for paragraph in doc.paragraphs:
                content += paragraph.text + "\n"
       
        elif file_name.endswith('.xlsx'):
            # Create a BytesIO object from the uploaded file
            excel_file = io.BytesIO(uploaded_file.read())
            try:
                workbook = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
               
                # Read all sheets
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    content += f"\n=== Sheet: {sheet_name} ===\n"
                   
                    # Read first 100 rows to avoid memory issues with large files
                    row_count = 0
                    for row in sheet.iter_rows(values_only=True, max_row=100):
                        row_content = [str(cell) if cell is not None else '' for cell in row]
                        content += "\t".join(row_content) + "\n"
                        row_count += 1
                        if row_count >= 100:  # Limit to 100 rows per sheet
                            content += f"\n... (showing first 100 rows only)\n"
                            break
            except Exception as e:
                print(f"Error reading XLSX file: {e}")
                content = f"Error reading Excel file: {str(e)}"
        
        elif file_name.endswith('.xls'):
            # Handle older Excel format (.xls) using xlrd
            excel_file = io.BytesIO(uploaded_file.read())
            try:
                workbook = xlrd.open_workbook(file_contents=excel_file.getvalue())
                
                # Read all sheets
                for sheet_index in range(min(workbook.nsheets, 5)):  # Limit to 5 sheets
                    sheet = workbook.sheet_by_index(sheet_index)
                    sheet_name = workbook.sheet_names()[sheet_index]
                    content += f"\n=== Sheet: {sheet_name} ===\n"
                    
                    # Read first 100 rows to avoid memory issues
                    for row_num in range(min(sheet.nrows, 100)):
                        row_values = []
                        for col_num in range(min(sheet.ncols, 50)):  # Limit to 50 columns
                            cell_value = sheet.cell_value(row_num, col_num)
                            row_values.append(str(cell_value) if cell_value != '' else '')
                        content += "\t".join(row_values) + "\n"
                    
                    if sheet.nrows > 100:
                        content += f"\n... (showing first 100 rows only)\n"
            except Exception as e:
                print(f"Error reading XLS file: {e}")
                content = f"Error reading Excel file: {str(e)}"
       
        else:
            return None
           
        return content.strip()
       
    except Exception as e:
        print(f"Error reading document: {e}")
        return None

def ask_gemini(document_content, user_question):
    """
    Send question to Gemini AI with document content
    """
    prompt_parts = [
        "You are a helpful document assistant. If the document contains photo then explain about the photo. Based on the content of the document provided, answer the user's question. You can convert any document into any language asked by the user.",
        "Document Content:\n" + document_content,
        "User Question:\n" + user_question,
    ]
   
    try:
        response = model.generate_content(prompt_parts)
        return response.text
    except Exception as e:
        print(f"Error communicating with Gemini: {e}")
        return "Sorry, I encountered an error while processing your question."